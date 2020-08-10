from openpyxl import load_workbook
import numpy as np
from scipy import stats
import pandas as pd
import scipy
import statsmodels.api as sm
wb = load_workbook('Data_files_to_richards/Table1_05042020_raw_P-values.xlsx')
print(scipy.__version__)
# print(wb.sheetnames)

ws = wb[wb.sheetnames[0]]

cont_rows = range(3,10)
total = 3841
deceased = 313
live = 3528

col_datasets = [['C', 'D', 'E', 'F'], ['H', 'I','J','K'], ['M', 'N', 'O','P']]

var_dict = {'AGE, mean (SD), years': 'AGE',
'TEMPERATURE at presentation, mean (SD), degrees Fahrenheit':'TEMPERATURE',
'TEMP_MAX during encounter, mean (SD), degrees Fahrenheit': 'TEMP_MAX',
'SYSTOLIC_BP at presentation, mean (SD), mm Hg': 'SYSTOLIC_BP',
'DIASTOLIC_BP at presentation, mean (SD), mm Hg': 'DIASTOLIC_BP',
'O2_SAT at presentation, mean (SD), percent': 'O2_SAT',
'O2SAT_MIN during encounter, mean (SD), percent': 'O2SAT_MIN'}

deceased_ind_col = 'DECEASED_INDICATOR'

dev_df = pd.read_csv('train_data_6th_April.csv')
print(dev_df.shape)
dev_deceased_df = dev_df[dev_df[deceased_ind_col]==1]
dev_live_df = dev_df[dev_df[deceased_ind_col]==0]

test1_df = pd.read_csv('test_data_6th_April.csv')
print(test1_df.shape)
test1_deceased_df = test1_df[test1_df[deceased_ind_col]==1]
test1_live_df = test1_df[test1_df[deceased_ind_col]==0]

test2_df = pd.read_csv('test_data_7th_April.csv')
print(test2_df.shape)
test2_deceased_df = test2_df[test2_df[deceased_ind_col]==1]
test2_live_df = test2_df[test2_df[deceased_ind_col]==0]

alpha = 0.05
z_crit = stats.norm.ppf(1-alpha/2)
print(z_crit)


def cohen_d(x,y):
    x = x[~np.isnan(x)]
    y = y[~np.isnan(y)]
    nx = len(x)
    ny = len(y)
    dof = nx + ny - 2
    d = (np.mean(x) - np.mean(y)) / np.sqrt(((nx-1)*np.std(x, ddof=1) ** 2 + (ny-1)*np.std(y, ddof=1) ** 2) / dof)
    se = np.sqrt((nx+ny)/(nx*ny)+(d**2)/(2*(nx+ny)))
    d_l = d-z_crit*se
    d_u = d+z_crit*se
    return d, d_l, d_u

def odds_ratio_logit(X, y):
    # y = y[~np.isnan(X)]
    # X = X[~np.isnan(X)]
    # X = X.fillna(X.mean())
    X['intercept'] = 1.0
    # print(X)
    # print(y)
    res = sm.Logit(list(y), X, missing='drop').fit()
    params = res.params
    conf = res.conf_int()
    conf['Odds Ratio'] = params
    conf.columns = ['5%', '95%', 'Odds Ratio']

    return np.exp(conf.values[0])
    # print(np.exp(conf))


# for cols in col_datasets:
for cont_r in cont_rows:
    # develop_total_mean = float(ws['B{}'.format(cont_r)].value.split(' (')[0])
    # develop_total_std = float(ws['B{}'.format(cont_r)].value.split(' (')[-1].split(')')[0])

    # print(develop_total_mean)
    cont_var = ws['A{}'.format(cont_r)].value
    var_in_df = var_dict[cont_var]
    dev_t_pval = stats.ttest_ind(dev_deceased_df[var_in_df].values, dev_live_df[var_in_df].values, equal_var=False, nan_policy='omit').pvalue
    test1_t_pval = stats.ttest_ind(test1_deceased_df[var_in_df].values, test1_live_df[var_in_df].values, equal_var=False, nan_policy='omit').pvalue
    test2_t_pval = stats.ttest_ind(test2_deceased_df[var_in_df].values, test2_live_df[var_in_df].values, equal_var=False, nan_policy='omit').pvalue

    # print(dev_t_pval)
    ws['{}{}'.format(col_datasets[0][2], cont_r)].value = dev_t_pval
    ws['{}{}'.format(col_datasets[1][2], cont_r)].value = test1_t_pval
    ws['{}{}'.format(col_datasets[2][2], cont_r)].value = test2_t_pval

    dev_or_logit = odds_ratio_logit(dev_df[[var_in_df]], dev_df[deceased_ind_col])
    test1_or_logit = odds_ratio_logit(test1_df[[var_in_df]], test1_df[deceased_ind_col])
    test2_or_logit = odds_ratio_logit(test2_df[[var_in_df]], test2_df[deceased_ind_col])




    # dev_cohen_d = cohen_d(dev_deceased_df[var_in_df].values, dev_live_df[var_in_df].values)
    # test1_cohen_d = cohen_d(test1_deceased_df[var_in_df].values, test1_live_df[var_in_df].values)
    # test2_cohen_d = cohen_d(test2_deceased_df[var_in_df].values, test2_live_df[var_in_df].values)

    ws['{}{}'.format(col_datasets[0][3], cont_r)].value = '{:.2f} ({:.2f},{:.2f})'.format(dev_or_logit[2], dev_or_logit[0], dev_or_logit[1])
    ws['{}{}'.format(col_datasets[1][3], cont_r)].value = '{:.2f} ({:.2f},{:.2f})'.format(test1_or_logit[2], test1_or_logit[0], test1_or_logit[1])
    ws['{}{}'.format(col_datasets[2][3], cont_r)].value = '{:.2f} ({:.2f},{:.2f})'.format(test2_or_logit[2], test2_or_logit[0], test2_or_logit[1])





cat_cols = [10,14,20,23,28,31,34,37,40,43,46]

cat_var_dict = {}

for cols in col_datasets:
    for idx, cat_start_col in enumerate(cat_cols):
        if idx < (len(cat_cols)-1):
            end_col = cat_cols[idx+1]
        else:
            end_col = 49
        cont_table = []
        cat_var_rows = range(cat_start_col+1, end_col)
        # print(cat_start_col)
        excel_row_dict = {}
        start_idx = 0
        for i in cat_var_rows:
            # print(i)
            deceased_count = int(ws['{}{}'.format(cols[0], i)].value.split(' (')[0])
            live_count = int(ws['{}{}'.format(cols[1], i)].value.split(' (')[0])
            if not(deceased_count == 0 and live_count == 0):
                excel_row_dict[start_idx] = i
                start_idx += 1
                cont_table.append([live_count, deceased_count])
        cont_table = np.array(cont_table)

        chi2_pval = stats.chi2_contingency(cont_table)[1]
        ws['{}{}'.format(cols[2], cat_start_col)].value = chi2_pval
        for idx_i, i in excel_row_dict.items():
            odds_ratio_table = np.zeros((2,2))
            for idx_j, j in excel_row_dict.items():
                if i == j:
                    odds_ratio_table[1,0] = cont_table[idx_i, 0]
                    odds_ratio_table[1,1] = cont_table[idx_i, 1]
                else:
                    odds_ratio_table[0,0] = odds_ratio_table[0,0] + cont_table[idx_j, 0]
                    odds_ratio_table[0,1] = odds_ratio_table[0,1] + cont_table[idx_j, 1]
            # print(odds_ratio_table)
            # if not (odds_ratio_table[1,0] == 0 and odds_ratio_table[1,1] == 0):
            table = sm.stats.Table2x2(odds_ratio_table)
            odds_ratio = table.oddsratio
            odds_ratio_ci = table.oddsratio_confint()
            odds_ratio_str = '{:.2f} ({:.2f},{:.2f})'.format(odds_ratio, odds_ratio_ci[0], odds_ratio_ci[1])
            # print('Odds ratio CI: ', odds_ratio_str)

            ws['{}{}'.format(cols[3],i)].value = odds_ratio_str


wb.save('document.xlsx')







# var_df = pd.read_csv('Data_files_to_richards/Training_data_6th_April_Richards.csv')
# deceased_ind_df = pd.read_csv('Data_files_to_richards/Training_label_6th_April_Richards.csv')
#
# #std deviation
# # s = np.sqrt((var_a + var_b)/2)
# ## Calculate the t-statistics
# # t = (a.mean() - b.mean())/(s*np.sqrt(2/N))
# s1 = (2.6**2)*deceased/(deceased-1)
# # s1 = 3.9**2
# s2 = (1.82**2)*live/(live-1)
#
# m1 = 101.11
# m2 = 100.25
# # s2 = 2.3
# s = np.sqrt(s1/deceased+s2/live)
# t = (m1-m2)/(s)
# df = deceased + live - 2
# p = 1 - stats.t.cdf(abs(t),df=df)
#
# print('t:', t)
# print('p_val:', 2*p)
#
# t_test_result = stats.ttest_ind_from_stats(mean1=m1, std1=np.sqrt(s1), nobs1=deceased-1,
#                                            mean2=m2, std2=np.sqrt(s2), nobs2=live-1,
#                                            equal_var=False)
#
#
#
#
#
# # calculate the pooled standard deviation
# pooled_s = np.sqrt(((deceased - 1) * s1 + (live - 1) * s2) / (deceased + live - 2))
# # calculate the means of the samples
# d = (m1-m2)/pooled_s
# print('d:', d)
#
# # d2 = t*np.sqrt(2*(1/deceased+1/live))
# d2 = t/np.sqrt(1/deceased+1/live)
# print('d:', d2)
#
#
# print(t_test_result)
#
# cat_df = pd.concat([var_df, deceased_ind_df], axis=1)
# deceased_df = cat_df[cat_df[deceased_ind_df.columns.values[0]]==1]
# live_df = cat_df[cat_df[deceased_ind_df.columns.values[0]]==0]
#
# m1 = deceased_df['TEMP_MAX'].mean()
# s1 = deceased_df['TEMP_MAX'].std(ddof=1)
# m2 = live_df['TEMP_MAX'].mean()
# s2 = live_df['TEMP_MAX'].std(ddof=1)
#
# print(m1,s1,m2,s2, len(deceased_df.index), len(live_df.index))
#
#
# t_test_result = stats.ttest_ind_from_stats(mean1=m1, std1=s1, nobs1=deceased,
#                                            mean2=m2, std2=s2, nobs2=live,
#                                            equal_var=True)
#
#
#
# print(t_test_result)
#
# t_test_result = stats.ttest_ind(deceased_df['TEMP_MAX'], live_df['TEMP_MAX'], equal_var=False)
#
#
# # obs = np.array([[191,1933,0],[121,1574,3]])
# # obs = np.array([[93,83,11,74,39],[915,890,151,858,489]])
# # obs = np.array([[6,307],[824,2704]])
#
# obs = np.array([[13,148,69,0],[121,1812,556,2]])
# chi2 = stats.chi2_contingency(obs.T)
# print(chi2)

# calculate the size of samples
# calculate the variance of the samples

# calculate the effect size